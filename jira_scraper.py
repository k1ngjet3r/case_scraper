'''
    This is primary for scrapping MY-22 cases from Jira
'''

from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook


class Scraper():
    def __init__(self, test_case_list, output_name):
        self.tc_list_sheet = (load_workbook(str(test_case_list))).active
        self.output_name = str(output_name)
        self.wb = Workbook()

    def case_list(self):
        id_list = []
        for id in self.tc_list_sheet.iter_rows(max_col=1, values_only=True):
            if id is not None:
                id_list.append(str(id)[2:-3])
            else:
                break
        return id_list

    def scrapping(self):
        print('Opening Jira...')
        try:
            driver = webdriver.Chrome()
            url = f'https://matsjira.cienetcorp.com/login.jsp'
            driver.get(url)
        except:
            raise SystemExit(
                'Unable to connect to Jira server, please check your wifi setting!')

        username = 'jeter.lin'
        pw = 'sD4T1pDTZp'
        print('Entering the username and password')
        driver.find_element_by_id('login-form-username').send_keys(username)
        driver.find_element_by_id('login-form-password').send_keys(pw)
        driver.find_element_by_id('login-form-submit').click()

        self.wb.active
        self.wb.create_sheet('Not found', 0)
        self.wb.create_sheet('Detailed list', 0)

        id_list = self.case_list()
        cur_num = 0
        total_tc = len(id_list)

        found_cases = 0
        not_found_cases = 0

        for id in id_list:
            cur_num += 1
            print('fatching...{}/{}'.format(cur_num, total_tc))
            driver.get(self.url_gen(id))
            try:
                original_TCID = driver.find_element_by_class_name(
                    'customfield_10202').text

                precondition = driver.find_element_by_class_name(
                    'customfield_10331').text
                test_step = driver.find_element_by_class_name(
                    'customfield_10342').text
                expected = driver.find_element_by_class_name(
                    'customfield_10315').text
                objective = driver.find_element_by_class_name(
                    'customfield_10336').text

                # summary = driver.find_element_by_class_name('issue-link').text
                # frop_1 = driver.find_element_by_class_name('customfield_10200').text
                # frop_2 = driver.find_element_by_class_name('customfield_10319').text

                case_detail = [original_TCID, precondition, test_step, expected, objective]
                found_cases += 1
                print('Found!')
                print('==========================================')
                self.wb['Detailed list'].append(case_detail)
            except:
                not_found_cases += 1
                print('cannot find the detail of case: {}'.format(id))
                print('==========================================')
                self.wb['Not found'].append([id])

        print('Done!, saving the file named {}'.format(self.output_name))
        self.wb.save(self.output_name)
        print('[SUMMARY]:')
        print('Found cases: {}'.format(found_cases))
        print('Not Found: {}'.format(not_found_cases))

    def url_gen(self, tcid):
        frame = f'https://matsjira.cienetcorp.com/issues/?jql=project%20%3D%20TESTSPEC22%20AND%20%22Original%20GM%20TC%20ID%22%20%20~%20'
        return frame + str(tcid)


scrp = Scraper('MY22_QIH22B-301_1200.xlsx', 'W15_301_cases.xlsx')

scrp.scrapping()
